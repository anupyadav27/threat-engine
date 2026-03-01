"""
Excel Exporter

Exports compliance reports as Excel (XLSX) for detailed analysis.
"""

from typing import Dict, List, Any, Optional
from io import BytesIO
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExporter:
    """Exports compliance reports as Excel (XLSX)."""
    
    @staticmethod
    def export_framework_report(report: Dict[str, Any]) -> bytes:
        """
        Export framework report as Excel with multiple sheets.
        
        Sheets:
        1. Summary - Overall framework statistics
        2. Controls - All controls with details
        3. Grouped by Control - Resources grouped by control ID
        4. Grouped by Resource - Controls grouped by resource
        
        Args:
            report: Framework report dictionary with grouping
        
        Returns:
            Excel file bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Summary
        summary_sheet = wb.create_sheet("Summary")
        ExcelExporter._add_summary_sheet(summary_sheet, report)
        
        # Sheet 2: Controls
        controls_sheet = wb.create_sheet("Controls")
        ExcelExporter._add_controls_sheet(controls_sheet, report)
        
        # Sheet 3: Grouped by Control
        if 'grouped_by_control' in report:
            control_group_sheet = wb.create_sheet("By Control")
            ExcelExporter._add_grouped_by_control_sheet(control_group_sheet, report['grouped_by_control'])
        
        # Sheet 4: Grouped by Resource
        if 'grouped_by_resource' in report:
            resource_group_sheet = wb.create_sheet("By Resource")
            ExcelExporter._add_grouped_by_resource_sheet(resource_group_sheet, report['grouped_by_resource'])
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def _add_summary_sheet(sheet, report: Dict[str, Any]):
        """Add summary sheet."""
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        sheet['A1'] = f"{report.get('framework', 'Compliance')} Report Summary"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:B1')
        
        row = 3
        # Summary statistics
        summary_data = [
            ['Framework', report.get('framework', '')],
            ['Compliance Score', f"{report.get('compliance_score', 0):.2f}%"],
            ['Overall Status', report.get('overall_status', '')],
            ['', ''],
            ['Statistics', ''],
            ['Total Controls', report.get('statistics', {}).get('controls_total', 0)],
            ['Controls Passed', report.get('statistics', {}).get('controls_passed', 0)],
            ['Controls Failed', report.get('statistics', {}).get('controls_failed', 0)],
            ['Controls Partial', report.get('statistics', {}).get('controls_partial', 0)],
        ]
        
        for data in summary_data:
            sheet[f'A{row}'] = data[0]
            sheet[f'B{row}'] = data[1] if len(data) > 1 else ''
            if data[0] == 'Statistics':
                sheet[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
        
        # Auto-adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30
    
    @staticmethod
    def _add_controls_sheet(sheet, report: Dict[str, Any]):
        """Add controls sheet."""
        # Header
        headers = ['Control ID', 'Control Title', 'Category', 'Status', 
                  'Total Checks', 'Passed', 'Failed', 'Compliance %']
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        controls = report.get('controls', [])
        for row_idx, control in enumerate(controls, 2):
            sheet.cell(row=row_idx, column=1, value=control.get('control_id', ''))
            sheet.cell(row=row_idx, column=2, value=control.get('control_title', ''))
            sheet.cell(row=row_idx, column=3, value=control.get('control_category', ''))
            sheet.cell(row=row_idx, column=4, value=control.get('status', ''))
            sheet.cell(row=row_idx, column=5, value=control.get('checks_total', 0))
            sheet.cell(row=row_idx, column=6, value=control.get('checks_passed', 0))
            sheet.cell(row=row_idx, column=7, value=control.get('checks_failed', 0))
            
            compliance_pct = control.get('compliance_percentage', 0)
            sheet.cell(row=row_idx, column=8, value=f"{compliance_pct:.2f}%")
            
            # Color code by status
            if control.get('status') == 'PASS':
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif control.get('status') == 'FAIL':
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            sheet.cell(row=row_idx, column=4).fill = fill
        
        # Auto-adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 15
        for col in range(5, 9):
            sheet.column_dimensions[get_column_letter(col)].width = 15
    
    @staticmethod
    def _add_grouped_by_control_sheet(sheet, grouped_data: Dict[str, Any]):
        """Add grouped by control sheet."""
        # Header
        headers = ['Control ID', 'Control Title', 'Resource ARN', 'Resource Type', 
                  'Region', 'Service', 'Status', 'Severity', 'Rule ID']
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 2
        for control_id, control_data in grouped_data.items():
            # Control header row
            sheet.cell(row=row, column=1, value=f"Control: {control_id}")
            sheet.cell(row=row, column=2, value=control_data.get('control_title', ''))
            sheet.merge_cells(f'A{row}:I{row}')
            sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
            sheet.cell(row=row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            # Failed resources
            for resource in control_data.get('resources_failed', []):
                sheet.cell(row=row, column=1, value=control_id)
                sheet.cell(row=row, column=2, value=control_data.get('control_title', ''))
                sheet.cell(row=row, column=3, value=resource.get('resource_arn', ''))
                sheet.cell(row=row, column=4, value=resource.get('resource_type', ''))
                sheet.cell(row=row, column=5, value=resource.get('region', ''))
                sheet.cell(row=row, column=6, value=resource.get('service', ''))
                sheet.cell(row=row, column=7, value='FAIL')
                sheet.cell(row=row, column=8, value=resource.get('severity', ''))
                sheet.cell(row=row, column=9, value=resource.get('rule_id', ''))
                sheet.cell(row=row, column=7).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                row += 1
            
            # Passed resources
            for resource in control_data.get('resources_passed', []):
                sheet.cell(row=row, column=1, value=control_id)
                sheet.cell(row=row, column=2, value=control_data.get('control_title', ''))
                sheet.cell(row=row, column=3, value=resource.get('resource_arn', ''))
                sheet.cell(row=row, column=4, value=resource.get('resource_type', ''))
                sheet.cell(row=row, column=5, value=resource.get('region', ''))
                sheet.cell(row=row, column=6, value=resource.get('service', ''))
                sheet.cell(row=row, column=7, value='PASS')
                sheet.cell(row=row, column=8, value=resource.get('severity', ''))
                sheet.cell(row=row, column=9, value=resource.get('rule_id', ''))
                sheet.cell(row=row, column=7).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                row += 1
            
            row += 1  # Space between controls
        
        # Auto-adjust column widths
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 20
    
    @staticmethod
    def _add_grouped_by_resource_sheet(sheet, grouped_data: Dict[str, Any]):
        """Add grouped by resource sheet."""
        # Header
        headers = ['Resource ARN', 'Resource Type', 'Region', 'Service', 
                  'Compliance Score', 'Control ID', 'Control Title', 'Status', 'Rule ID']
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 2
        for resource_arn, resource_data in grouped_data.items():
            # Resource header row
            resource_info = resource_data.get('resource_info', {})
            sheet.cell(row=row, column=1, value=f"Resource: {resource_arn}")
            sheet.cell(row=row, column=5, value=f"{resource_data.get('compliance_score', 0):.2f}%")
            sheet.merge_cells(f'A{row}:I{row}')
            sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
            sheet.cell(row=row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            # Failed controls
            for control in resource_data.get('failed_controls', []):
                for check in control.get('checks', []):
                    sheet.cell(row=row, column=1, value=resource_arn)
                    sheet.cell(row=row, column=2, value=resource_info.get('resource_type', ''))
                    sheet.cell(row=row, column=3, value=resource_info.get('region', ''))
                    sheet.cell(row=row, column=4, value=resource_info.get('service', ''))
                    sheet.cell(row=row, column=5, value=f"{resource_data.get('compliance_score', 0):.2f}%")
                    sheet.cell(row=row, column=6, value=control.get('control_id', ''))
                    sheet.cell(row=row, column=7, value=control.get('control_title', ''))
                    sheet.cell(row=row, column=8, value='FAIL')
                    sheet.cell(row=row, column=9, value=check.get('rule_id', ''))
                    sheet.cell(row=row, column=8).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    row += 1
            
            # Passed controls
            for control in resource_data.get('passed_controls', []):
                for check in control.get('checks', []):
                    sheet.cell(row=row, column=1, value=resource_arn)
                    sheet.cell(row=row, column=2, value=resource_info.get('resource_type', ''))
                    sheet.cell(row=row, column=3, value=resource_info.get('region', ''))
                    sheet.cell(row=row, column=4, value=resource_info.get('service', ''))
                    sheet.cell(row=row, column=5, value=f"{resource_data.get('compliance_score', 0):.2f}%")
                    sheet.cell(row=row, column=6, value=control.get('control_id', ''))
                    sheet.cell(row=row, column=7, value=control.get('control_title', ''))
                    sheet.cell(row=row, column=8, value='PASS')
                    sheet.cell(row=row, column=9, value=check.get('rule_id', ''))
                    sheet.cell(row=row, column=8).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    row += 1
            
            row += 1  # Space between resources
        
        # Auto-adjust column widths
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 25
    
    @staticmethod
    def export_executive_summary(report: Dict[str, Any]) -> bytes:
        """
        Export executive dashboard as Excel.
        
        Args:
            report: Executive dashboard dictionary
        
        Returns:
            Excel file bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        
        # Title
        ws['A1'] = "Compliance Executive Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Summary statistics
        summary = report.get('summary', {})
        row = 3
        ws.cell(row=row, column=1, value="Overall Compliance Score")
        ws.cell(row=row, column=2, value=f"{summary.get('overall_compliance_score', 0):.2f}%")
        row += 1
        
        ws.cell(row=row, column=1, value="Total Frameworks")
        ws.cell(row=row, column=2, value=summary.get('total_frameworks', 0))
        row += 2
        
        # Framework table header
        headers = ['Framework', 'Compliance Score', 'Status', 'Controls Total', 
                  'Controls Passed', 'Controls Failed']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        # Framework data
        frameworks = report.get('frameworks', [])
        for fw in frameworks:
            ws.cell(row=row, column=1, value=fw.get('framework', ''))
            ws.cell(row=row, column=2, value=f"{fw.get('compliance_score', 0):.2f}%")
            ws.cell(row=row, column=3, value=fw.get('status', ''))
            ws.cell(row=row, column=4, value=fw.get('controls_total', 0))
            ws.cell(row=row, column=5, value=fw.get('controls_passed', 0))
            ws.cell(row=row, column=6, value=fw.get('controls_failed', 0))
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
